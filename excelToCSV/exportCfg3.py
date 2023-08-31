import os
import logging
import csv
import openpyxl
import luaCodeGen
import argparse
from luaCodeGen import LuaCodeGen

INPUTPATH = u"excel_new"
SERVERPATH = u"csv_server"
CLIENTPATH = u"csv_client"
LUACSVOUTPATH = u"csv_lua"
ExportLuaFileCfgPath = u"exportLuaFileCfg.txt"
ExportCSVFileCfgPath = u"exportCSVFileCfg.txt"
SplitCol = 3
NameCol = 2
Col = 4
IsUseExportList = False


class changeLuaCenter:
    def __init__(self):
        self.mData = None

    def getvalue(self, filename):
        self.mData = []
        workbook = openpyxl.load_workbook(filename, data_only=True)
        table = workbook.worksheets[0]
        max_row = table.max_row
        max_column = table.max_column
        for i in range(1, max_row + 1):
            row = []
            for j in range(1, max_column + 1):
                value = table.cell(i, j).value
                if value is not None:
                    value = str(value).strip()
                split_type = table.cell(SplitCol, j).value
                nameKey = table.cell(NameCol,j).value
                if nameKey is not None and split_type is not None:                   
                    if 'a' in split_type or 'l' in split_type:
                        row.append(value)
            if i != SplitCol:
                self.mData.append(tuple(row))

    def write(self, path, filename):
        if not os.path.exists(path):
            os.makedirs(path)
        with open("tmp", "w", newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(self.mData)
        csvfile.close()

        if os.path.exists(os.path.join(path, filename)):
            os.remove(os.path.join(path, filename))
        os.rename('tmp', os.path.join(path, filename))
        logging.info("write file finish")
        print("write", filename, " finish")


class changeCSVCenter:
    def __init__(self):
        self.mData = None

    def getvalue(self, filename):
        self.mData = []
        print(filename + "filename:")
        workbook = openpyxl.load_workbook(filename, data_only=True)
        table = workbook.worksheets[0]
        max_row = table.max_row
        max_column = table.max_column
        for i in range(1, max_row + 1):
            row = []
            for j in range(1, max_column + 1):
                value = table.cell(i, j).value
                if value is not None:
                    value = str(value).strip()
                splitType = table.cell(SplitCol, j).value
                nameKey = table.cell(NameCol, j).value
                if nameKey is not None and splitType is not None:
                    if 'a' in splitType or 'c' in splitType:
                        row.append(value)

            if i != SplitCol:
                self.mData.append(tuple(row))

    def write(self, path, filename):
        if not os.path.exists(path):
            os.makedirs(path)

        with open("tmp", "w", newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(self.mData)
        csvfile.close()

        if os.path.exists(os.path.join(path, filename)):
            os.remove(os.path.join(path, filename))
        os.rename('tmp', os.path.join(path, filename))
        logging.info("write file finish")
        print("write", filename, " finish")


class changeServerCenter:
    def __init__(self):
        self.mData = None

    def getvalue(self, filename):
        self.mData = []
        workbook = openpyxl.load_workbook(filename, data_only=True)
        table = workbook.worksheets[0]
        max_row = table.max_row
        max_column = table.max_column
        for i in range(1, max_row + 1):
            row = []
            for j in range(1, max_column + 1):
                value = table.cell(i, j).value
                if value is not None:
                    value = str(value).strip()
                splitType = table.cell(SplitCol, j).value
                nameKey = table.cell(NameCol, j).value
                if nameKey is not None:
                    if splitType is not None:
                        if 'a' in splitType or 's' in splitType or splitType == '':
                            row.append(value)
                    else:
                        row.append(value)

            if i != SplitCol:
                self.mData.append(tuple(row))

    def write(self, path, filename):
        if not os.path.exists(path):
            os.makedirs(path)
        with open("tmp", "w", newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(self.mData)
        csvfile.close()

        if os.path.exists(os.path.join(path, filename)):
            os.remove(os.path.join(path, filename))
        os.rename('tmp', os.path.join(path, filename))
        logging.info("write file finish")
        print("write", filename, " finish")


def handleClear():
    clearOldFile(SERVERPATH)
    clearOldFile(CLIENTPATH)
    clearOldFile(LUACSVOUTPATH)


def manuallyBuild(paths):
    for path in paths:
        if os.path.exists(path):
            print(path)
            st = changeCSVCenter()
            st.getvalue(path)
            name = path.split("/")[-1].split(".")[0]
            st.write(CLIENTPATH, name + ".csv")

            st = changeLuaCenter()
            st.getvalue(path)
            name = path.split("/")[-1].split(".")[0]
            st.write(LUACSVOUTPATH, name + ".csv")

            st = changeServerCenter()
            st.getvalue(path)
            name = path.split("/")[-1].split(".")[0]
            st.write(SERVERPATH, name + ".csv")
        else:
            print(path + " don't exist")
            return


def handleExcel():
    handleClear()
    csvCfgFiles = []
    luaCfgFiles = []
    if IsUseExportList:
        csvCfgFile = open(ExportCSVFileCfgPath, 'r')
        line = csvCfgFile.readline()
        while line:
            csvCfgFiles.append(line.rstrip())
            line = csvCfgFile.readline()
        csvCfgFile.close()

        luaCfgFile = open(ExportLuaFileCfgPath, 'r')
        line = luaCfgFile.readline()
        while line:
            luaCfgFiles.append(line.rstrip())
            line = luaCfgFile.readline()
        luaCfgFile.close()

    files, dirs, root = readFilename(INPUTPATH)
    for fi in files:
        if fi in csvCfgFiles or len(csvCfgFiles) == 0 or not IsUseExportList:
            print("export excel " + fi + " to " + CLIENTPATH)
            str_stock = os.path.join(INPUTPATH, fi)
            if os.path.exists(str_stock):
                st = changeCSVCenter()
                st.getvalue(str_stock)
                name = fi.replace(".xlsx", "")
                st.write(CLIENTPATH, name + ".csv")
            else:
                print(str_stock + " don't exist")

        if fi in luaCfgFiles or len(luaCfgFiles) == 0 or not IsUseExportList:
            print("export excel " + fi + " to " + LUACSVOUTPATH)
            str_stock = os.path.join(INPUTPATH, fi)
            if os.path.exists(str_stock):
                st = changeLuaCenter()
                st.getvalue(str_stock)
                name = fi.replace(".xlsx", "")
                st.write(LUACSVOUTPATH, name + ".csv")
            else:
                print(str_stock + " don't exist")

        print("export excel " + fi + " to " + SERVERPATH)
        str_stock222 = os.path.join(INPUTPATH, fi)
        if os.path.exists(str_stock222):
            st22 = changeServerCenter()
            st22.getvalue(str_stock222)
            name = fi.replace(".xlsx", "")
            st22.write(SERVERPATH, name + ".csv")
        else:
            print(str_stock222 + " don't exist")


def readFilename(file_dir):
    for root, dirs, files in os.walk(file_dir):
        return files, dirs, root


def clearOldFile(file_dir):
    if not os.path.exists(file_dir):
        os.makedirs(file_dir)

    files, dirs, root = readFilename(file_dir)
    for fi in files:
        strstock = os.path.join(file_dir, fi)
        if os.path.exists(strstock):
            os.remove(strstock)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="can input -all -o -auto")
    parser.add_argument("-all", "--buildAll", action="store_true", help="build all excel mode,folder is excel")
    parser.add_argument("-auto", "--autoBuild", action="store_true", help="auto build all excel in folder is excel_new")
    parser.add_argument("-o", "--output", nargs='+', help="build the spec excel mode,enter param manually,like "
                                                          "./excel/aaa.xlsx")

    args = parser.parse_args()
    if args.buildAll:
        INPUTPATH = u"excel"
        IsUseExportList = True
        handleExcel()
        luaCodeGen.LuaCodeGen.GenLuaCode()
    elif args.autoBuild:
        INPUTPATH = u"excel_new"
        IsUseExportList = False
        handleExcel()
        luaCodeGen.LuaCodeGen.GenLuaCode()
    elif args.output:
        manuallyBuild(args.output)
        luaCodeGen.LuaCodeGen.GenLuaCode()
    else:
        INPUTPATH = u"excel_new"
        IsUseExportList = False
        handleExcel()
        luaCodeGen.LuaCodeGen.GenLuaCode()
