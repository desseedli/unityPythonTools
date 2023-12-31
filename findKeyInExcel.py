import openpyxl
import os

file_path = './excel'
find_list = ['rmb', 'RMB']
res_path = []

file_list = os.listdir(file_path)
for file_name in file_list:
    if file_name.endswith('xlsx'):
        file_name_no_extension, extension = os.path.splitext(file_name)
        open_file_path = os.path.join(file_path, file_name)
        workbook = openpyxl.load_workbook(open_file_path, data_only=True)
        print('open ' + file_name_no_extension)
        table = workbook.worksheets[0]
        max_column = table.max_column
        read_row = 2
        for i in range(1, max_column + 1):
            value = table.cell(2, i).value
            if value in find_list:
                res_path.append(file_name_no_extension)
                #print(file_name_no_extension)
                break

for res in res_path:
    print(res)
