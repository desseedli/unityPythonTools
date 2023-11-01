import os
import logging
import csv
import openpyxl
import luaCodeGen
import argparse
import sys
import time
from io import StringIO
from luaCodeGen import LuaCodeGen

INPUTPATH = u"excel"
SERVERPATH = u"csv_server"
CLIENTPATH = u"csv_client"
LUAPATH = u"lua"

LOCALIZATIONPATH=u"..\\Assets\\Content\\Localization\\CSV"
LOCALIZATIONFILE = u"excel_loc\\Localization.xlsx"
PATCHLOCALIZATIONPATH=u"..\\Assets\\Content\\PatchSystem\\Localization"
PATCHLOCALIZATIONKEY="LOC_PATCHING_"
ExportLuaFileCfgPath = u"exportLuaFileCfg.txt"
ExportCSVFileCfgPath = u"exportCSVFileCfg.txt"
IsUseExportList = False




# ANSI escape sequence for red color
RED = "\033[91m"
GREEN = "\033[92m"
YELLOW = "\033[93m"
DARK = "\033[90m"
RESET = "\033[0m"

TIMER = 0

ExportMode = {'Client':0, 'Server':1, 'Lua':2}

def print_dim(message):
    print(DARK +  message + RESET)

def print_error(message):
    print(RED +  message + RESET)

def print_green(message):
    print(GREEN + message + RESET)

def print_warning(message):
    print(YELLOW + message + RESET)

def debug_start_time():
    global TIMER
    TIMER = time.time();

def debug_get_time():
    global TIMER
    return time.time() - TIMER


class csvExporter:
    def __init__(self):
        self.mData = [[],[],[]]
        self.mColumnProperties = []

    def getvalue(self, filename):
        workbook = openpyxl.load_workbook(filename, data_only=True)
        table = workbook.worksheets[0]
        max_row = table.max_row
        max_column = table.max_column

        #init column properties
        SplitRow = 3
        NameRow = 2
        for i in range(1, max_column + 1):
            split_type = table.cell(SplitRow, i).value
            nameKey = table.cell(NameRow, i).value
            if nameKey is not None and split_type is not None:
                self.mColumnProperties.append(split_type)
            elif nameKey is not None and split_type is None:
                self.mColumnProperties.append("")
            else:
                self.mColumnProperties.append(None)
        for i in range(1, max_row + 1):
            row = [[],[],[]]
            for j in range(1, max_column + 1):
                columnProperty = self.mColumnProperties[j-1]
                if columnProperty != None:                
                    value = table.cell(i, j).value
                    if value is not None:
                        value = str(value).strip()
                        if "float" == table.cell(4, j).value and value.lstrip("-").isnumeric():
                            value += ".0"
                        if value == "#N/A": # fix #N/A default to 42
                            value = "42"
                    if "bool" == table.cell(4, j).value and value is None:
                        value = "false"
                    if j == 1 and value is None:
                        break
                    if 'a' in columnProperty:
                        for k in range(len(ExportMode)):
                            row[k].append(value)
                    else:
                        if 'c' in columnProperty:
                            row[ExportMode['Client']].append(value)
                        if 'l' in columnProperty:
                            row[ExportMode['Lua']].append(value)
                        if 's' in columnProperty or columnProperty == "":
                            row[ExportMode['Server']].append(value)
            if i != SplitRow and None != table.cell(i, 1).value:
                for k in range(len(ExportMode)):
                    self.mData[k].append(tuple(row[k]))

    def write(self, path, filename, mode):
        if not os.path.exists(path):
            os.makedirs(path)
        filePath = os.path.join(path, filename)
        with open(filePath, "w", newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(self.mData[mode])
        csvfile.close()
        return True

    def write_memory(self, mode):
        memory_file = StringIO()
        writer = csv.writer(memory_file)
        writer.writerows(self.mData[mode])
        memory_file.seek(0)
        return memory_file

def handleClear():
    clearOldFile(SERVERPATH)
    clearOldFile(CLIENTPATH)
    clearOldFile(LUAPATH)

def manuallyBuild(paths):
    for path in paths:
        processExcelFile(path, [], [])

def processExcelFile(filePath, csvCfgFiles, luaCfgFiles):
    if not os.path.exists(filePath):
        print_error("Error: " + filePath + " doesn't exist")
        return

    debug_start_time()
    fileName = os.path.basename(filePath)
    print("Exporting: " + filePath)
    
    st = csvExporter()
    st.getvalue(filePath)

    baseName =  os.path.splitext(fileName)[0]
    if fileName in csvCfgFiles or len(csvCfgFiles) == 0 or not IsUseExportList:
        if st.write(CLIENTPATH, baseName + ".csv", ExportMode['Client']):
            print_green("- CSV Client: {}".format(os.path.join(CLIENTPATH, baseName + ".csv")))
        else:
            print_error("- CSV Client: " + "write " + os.path.join(CLIENTPATH, baseName + ".csv") + " failed")
    elif IsUseExportList:
            print_warning("- CSV Client: Excluding by configuration")

    if st.write(SERVERPATH, baseName + ".csv", ExportMode['Server']):
        print_green("- CSV Server: {}".format(os.path.join(SERVERPATH, baseName + ".csv")))
    else:
        print_error("- CSV Server: Write " + os.path.join(SERVERPATH, baseName + ".csv") + " failed")

    if fileName in luaCfgFiles or len(luaCfgFiles) == 0 or not IsUseExportList:
        memory_file = st.write_memory(ExportMode['Lua'])
        if memory_file:
            luaCodeGen = LuaCodeGen()
            luafiles = luaCodeGen.gen_lua_code(memory_file, os.path.join(LUAPATH, baseName + ".lua"))            
            for luafile in luafiles:
                print_green("- Lua       : {}".format(luafile))
        else:
            print_error("- Lua       : Write lua file failed")
    elif IsUseExportList:
            print_warning("- Lua       : Excluding by configuration")

    print_dim("Completed in {:.2f} seconds.".format(debug_get_time())) 
    print("")

def handleExcel():
    total_timer = time.time();
    handleClear()
    csvCfgFiles = []
    luaCfgFiles = []
    if IsUseExportList:
        csvCfgFile = open(ExportCSVFileCfgPath, 'r')
        line = csvCfgFile.readline()
        while line:
            csvCfgFiles.append(line.rstrip())
            line = csvCfgFile.readline()
        csvCfgFile.close()

        luaCfgFile = open(ExportLuaFileCfgPath, 'r')
        line = luaCfgFile.readline()
        while line:
            luaCfgFiles.append(line.rstrip())
            line = luaCfgFile.readline()
        luaCfgFile.close()

    files, dirs, root = readFilename(INPUTPATH)

    for file in [ fi for fi in files if fi.endswith(".xlsx") and not fi.startswith("~") ]:
        processExcelFile(os.path.join(INPUTPATH, file), csvCfgFiles, luaCfgFiles)

    print("Total time taken: {:.2f} seconds.".format(time.time() - total_timer))

def readFilename(file_dir):
    for root, dirs, files in os.walk(file_dir):
        return files, dirs, root


def clearOldFile(file_dir):
    print("Clean folder " + file_dir)
    if not os.path.exists(file_dir):
        os.makedirs(file_dir)

    files, dirs, root = readFilename(file_dir)
    for fi in files:
        strstock = os.path.join(file_dir, fi)
        if os.path.exists(strstock):
            os.remove(strstock)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="can input -all -f -auto")
    parser.add_argument("-x", "--excelPath", help="excel source directory path")
    parser.add_argument("-all", "--buildAll", action="store_true", help="build all excel file folder is excel")
    parser.add_argument("-auto", "--autoBuild", action="store_true", help="auto build all excel in folder is excel_new")
    parser.add_argument("-loc", "--buildLocalization", action="store_true", help="build localization")
    parser.add_argument("-o", "--output", nargs='+', help="build the spec excel file mode, enter param manually like "
                                                          "./excel/GlobalCfg.xlsx")
    args = parser.parse_args()
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(1)

    if args.excelPath:
        INPUTPATH = args.excelPath

    if not os.path.exists(INPUTPATH):
        print_error("Error: The path '" + INPUTPATH + "' does not exist.")
        sys.exit(1)


    if args.buildLocalization:
        if not os.path.exists(LOCALIZATIONPATH):
            os.makedirs(LOCALIZATIONPATH)
        if not os.path.exists(PATCHLOCALIZATIONPATH):
            os.makedirs(PATCHLOCALIZATIONPATH)
        print("Building localization file " + LOCALIZATIONFILE)
        workbook = openpyxl.load_workbook(LOCALIZATIONFILE, data_only=True)
        table = workbook.worksheets[0]
        max_row = table.max_row
        max_column = table.max_column
        duplicate_key_found = False
        for cl in range(2, max_column + 1):
            key = {}
            data = []
            patchData = []
            for i in range(4, max_row + 1): # start from 4 row to skip the header
                row = []
                k = table.cell(i, 1).value
                check_duplicate_key = (cl == 2)
                if check_duplicate_key:
                    getkey = key.get(k)
                    if not getkey:
                        key[k] = [i]
                    else:
                        key[k].append(i)
                row.append(table.cell(i, 1).value)
                row.append(table.cell(i, cl).value)
                if type(k) == str and  k.startswith(PATCHLOCALIZATIONKEY):
                    patchData.append(row)
                data.append(row)

            if check_duplicate_key:
                for k, lines in key.items():
                    if len(lines) > 1:
                        duplicate_key_found = True
                        print_warning("Warning: Duplicated key '" + k + "' found at row " + str(lines))
            if not duplicate_key_found or True:
                lang = table.cell(2, cl).value
                filePath = os.path.join(LOCALIZATIONPATH, os.path.splitext(os.path.basename(LOCALIZATIONFILE))[0] + table.cell(2, cl).value + ".csv")
                print("- {:<10} : {}".format(lang + " Game", filePath))
                with open(filePath, "w", newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerows(data)
                csvfile.close()

                filePath = os.path.join(PATCHLOCALIZATIONPATH, os.path.splitext(os.path.basename(LOCALIZATIONFILE))[0] + table.cell(2, cl).value + ".csv")
                print("- {:<10} : {}".format(lang + " Patch", filePath))
                with open(filePath, "w", newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerows(patchData)
                csvfile.close()

        if duplicate_key_found:
            print_warning("Build localization with duplicate key found, please check the output.")
        else:
            print_green("Build localization success");            
    elif args.buildAll:
        IsUseExportList = True
        handleExcel()
    elif args.autoBuild:
        INPUTPATH = u"excel_new"
        IsUseExportList = False
        handleExcel()
    elif args.output:
        manuallyBuild(args.output)
    else:
        print_error("Can not handle")
