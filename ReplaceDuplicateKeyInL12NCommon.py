import os
import importlib
import subprocess
import csv


def check_and_install_module(module_name):
    try:
        importlib.import_module(module_name)
        print(f"模块 {module_name} 已安装")
    except ImportError:
        print(f"模块 {module_name} 未安装，正在安装...")
        try:
            subprocess.check_call(['pip', 'install', module_name])
            print(f"模块 {module_name} 安装成功")
        except subprocess.CalledProcessError:
            print(f"模块 {module_name} 安装失败")


need_check_module = True
if need_check_module:
    check_and_install_module("openpyxl")

from openpyxl import load_workbook

# pip install openpyxl
# pip uninstall openpyxl

open_path = os.path.abspath("../Tools/replaceExcel/duplicateL12N.txt")
walk_folder_path = os.path.abspath("../../../Tool/excel")
save_xlsx_folder_path = os.path.abspath("../Tools/replaceExcel")
koko_read_csv_path = os.path.abspath("../Tools/replaceExcel/LocalizationCommon.csv")

os.makedirs(save_xlsx_folder_path, exist_ok=True)

print(walk_folder_path)

res_data = []
need_check_false_data = []
with open(open_path, "r", encoding='utf-8') as file:
    for line in file:
        line = line.strip()
        if line:
            temp = eval(line)
            if temp['need_check'] == "true":
                res_data.append(temp)
            else:
                need_check_false_data.append(temp)


def replace_key(cell):
    for data in res_data:
        if data['key'] == cell.value:
            print("Have Not Separator:", data['key'], data['replace_key'])
            cell.value = data['replace_key']


def replace_key_in_list(cell, value_list):
    is_found = False
    for i in range(len(value_list)):
        for data in res_data:
            temp_str = value_list[i]
            if temp_str == data['key']:
                print("Have Separator:", "original key:", data['key'], "replace_key:", data['replace_key'])
                value_list[i] = data['replace_key']
                is_found = True

    if is_found:
        res = ";".join(value_list)
        print("Cells Data:", res)
        cell.value = res


file_list = os.listdir(walk_folder_path)
for file_name in file_list:
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(walk_folder_path, file_name)
        print("open file:" + file_name)
        workbook = load_workbook(filename=file_path, data_only=True)
        sheet = workbook.active
        skip_rows = 5
        for row in range(skip_rows, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, column)
                if cell.value is not None:
                    value = cell.value
                    if isinstance(value, str):
                        value_list = value.split(";")
                        if len(value_list) > 1:
                            replace_key_in_list(cell, value_list)
                        else:
                            replace_key(cell)

        target_file_path = os.path.join(save_xlsx_folder_path, file_name)
        workbook.save(filename=target_file_path)

with open(koko_read_csv_path, 'r', encoding='utf-8') as file:
    reader = csv.reader(file)
    rows = list(reader)

csv_new_row_data = []
for key_value in need_check_false_data:
    temp_data = []
    for key, value in key_value.items():
        if key != "replace_key" or key != "need_check":
            temp_data.append(value)
    csv_new_row_data.append(temp_data)

for new_row in csv_new_row_data:
    rows.append(new_row)

with open(koko_read_csv_path, 'w', encoding='utf-8', newline='') as file:
    writer = csv.writer(file)
    writer.writerows(rows)

print("done")
